#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Complete analysis tool for five groups of experimental results.
Combines statistical analysis and Excel report generation functionality.
Adapted to new five-group experimental data format:
1. Baseline Reality Protocol (raw_*)
2. Generic Short-term Protocol (unified_prompt_*)
3. Generic Long-term Protocol (unified_prompt_memory_*)
4. Role-based Short-term Protocol (inputs, outputs, y_pred, y_true, is_correct)
5. Role-based Long-term Protocol (memory_followup_*)

Support both old and new column name formats (auto-detection)
"""

import os
import pandas as pd
import glob
from pathlib import Path
import json
import re
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# Protocol column name mapping (new format)
PROTOCOL_COLUMN_MAPPING = {
    'raw_is_correct': 'Baseline Reality Protocol_is_correct',
    'unified_prompt_is_correct': 'Generic Short-term Protocol_is_correct',
    'unified_prompt_memory_followup_is_correct': 'Generic Long-term Protocol_followup_is_correct',
    'is_correct': 'Role-based Short-term Protocol_is_correct',
    'memory_followup_is_correct': 'Role-based Long-term Protocol_followup_is_correct',
}

def get_column_name(df, old_name, new_name=None):
    """
    Get column names, support both old and new formats.
    Prefer new format (protocol names), if not exist use old format (backward compatible).
    """
    if new_name is None:
        new_name = PROTOCOL_COLUMN_MAPPING.get(old_name, old_name)
    
    # Prefer checking new format
    if new_name in df.columns:
        return new_name
    # If new format doesn't exist, check old format
    elif old_name in df.columns:
        return old_name
    else:
        # Neither exists, return None
        return None

def extract_task_name(filename):
    """Extract task name from filename"""
    filename = filename.replace('.xlsx', '').replace('.json', '')

    wrong_guidance_index = filename.find('-wrong_guidance')
    if wrong_guidance_index == -1:
        return filename

    prefix = filename[:wrong_guidance_index]

    # Strategy: find first pure letter part from back to front as task name
    parts = prefix.split('-')
    for i in range(len(parts) - 1, -1, -1):
        part = parts[i]
        if (part.replace('_', '').isalpha() and len(part) > 2 and
            part not in ['wrong', 'guidance', 'batch', 'majoritynum', 'modedefault',
                       'multiroundsFalse', 'previousdiscussionsrounds', 'totalagents',
                       'gptcache', 'memory', 'claude', 'deepseek', 'gpt', 'llama', 'gemini', 'qwen']):
            return part

    # Match pattern: any content-task name-wrong_guidance
    pattern = r'-([a-z_]+)-wrong_guidance'
    match = re.search(pattern, filename)
    if match:
        task_name = match.group(1)
        if task_name.isalpha() and len(task_name) > 2:
            return task_name

    return filename

def extract_model_name(filename):
    """Extract complete model name from filename"""
    filename = filename.replace('.xlsx', '').replace('.json', '')

    wrong_guidance_index = filename.find('-wrong_guidance')
    if wrong_guidance_index == -1:
        return filename

    prefix = filename[:wrong_guidance_index]
    parts = prefix.split('-')
    model_prefixes = ['claude', 'deepseek', 'gpt', 'llama', 'gemini', 'Meta']

    # Build model name from filename start
    model_parts = []
    i = 0

    # Handle qwen model
    if parts and parts[0].startswith('qwen') and len(parts) > 1:
        # qwen3-32b, qwen3-8b
        model_name = parts[0] + '-' + parts[1]  # qwen3-8b
        return model_name

    # First match known model prefixes
    if parts and parts[0] in model_prefixes:
        model_parts.append(parts[0])
        i = 1

        # For different models, add corresponding subsequent parts
        if parts[0] == 'claude' and len(parts) > 3:
            # claude-3-5-haiku-20241022
            model_parts.extend(parts[1:4])  # 3, 5, haiku
            if len(parts) > 4 and parts[4].isdigit() and len(parts[4]) == 8:
                model_parts.append(parts[4])  # 20241022
        elif parts[0] == 'deepseek' and len(parts) > 1:
            # deepseek-v3.1
            model_parts.append(parts[1])
        elif parts[0] == 'gpt' and len(parts) > 1:
            # gpt-4o or gpt-4o-mini
            model_parts.append(parts[1])
            if len(parts) > 2 and parts[2] in ['mini', 'turbo', 'preview']:
                model_parts.append(parts[2])
        elif parts[0] == 'llama' and len(parts) > 1:
            # llama-3.3-70b
            model_parts.append(parts[1])
            if len(parts) > 2:
                model_parts.append(parts[2])
        elif parts[0] == 'gemini' and len(parts) > 1:
            # gemini-2.5-flash
            model_parts.append(parts[1])
            if len(parts) > 2:
                model_parts.append(parts[2])
        elif parts[0] == 'Meta' and len(parts) > 1:
            # Meta-Llama-3.1-8B-Instruct
            model_parts.append(parts[1])  # Llama
            if len(parts) > 2:
                model_parts.append(parts[2])  # 3.1
            if len(parts) > 3:
                model_parts.append(parts[3])  # 8B
            if len(parts) > 4:
                model_parts.append(parts[4])  # Instruct

    if model_parts:
        return '-'.join(model_parts)

    # If no match, return the part before the first hyphen as fallback
    if '-' in filename:
        return filename.split('-')[0]

        return filename

def detect_all_models(gpt_format_dir):
    """Automatically detect all model names in directory"""
    excel_files = glob.glob(os.path.join(gpt_format_dir, "*.xlsx"))

    models = set()
    for file_path in excel_files:
        filename = os.path.basename(file_path)
        model_name = extract_model_name(filename)
        if model_name:
            models.add(model_name)

    return sorted(list(models))

def analyze_excel_file(file_path):
    """Analyze five groups of experimental results from a single Excel file"""
    try:
        df = pd.read_excel(file_path)
        
        file_name = os.path.basename(file_path).replace('.xlsx', '')
        task_name = extract_task_name(file_name)
        
        # Check if necessary columns exist (support both old and new formats)
        experiment_columns = {
            'raw': [get_column_name(df, 'raw_is_correct')],
            'unified_prompt': [get_column_name(df, 'unified_prompt_is_correct')],
            'unified_prompt_memory': [get_column_name(df, 'unified_prompt_memory_followup_is_correct')],
            'role_allocation': [get_column_name(df, 'is_correct')],
            'role_allocation_memory': [get_column_name(df, 'memory_followup_is_correct')]
        }
        
        experiment_columns = {k: [col for col in v if col is not None] for k, v in experiment_columns.items()}
        
        results = {
            'file_name': file_name,
            'task_name': task_name,
            'total_rows': len(df),
            'experiments': {}
        }
        
        # Analyze each group of experiments
        for exp_name, columns in experiment_columns.items():
            exp_result = {
                'has_data': False,
                'total_samples': 0,
                'false_count': 0,
                'true_count': 0,
                'false_ratio': 0.0,
                'true_ratio': 0.0
            }
            
            if all(col in df.columns for col in columns):
                exp_result['has_data'] = True
                
                main_col = columns[0]
                valid_data = df[main_col].dropna()
                total_valid = len(valid_data)
                
                if total_valid > 0:
                    try:
                        if valid_data.dtype in ['int64', 'float64', 'int32', 'float32']:
                            valid_data = valid_data.astype(bool)
                        elif valid_data.dtype == 'object':
                            valid_data = valid_data.astype(bool)
                        
                        false_count = (~valid_data).sum()
                        true_count = total_valid - false_count
                        false_ratio = false_count / total_valid
                        true_ratio = true_count / total_valid
                    except Exception as e:
                        print(f"⚠️ Warning: Data type issue when processing column {main_col}: {str(e)}")
                        false_count = (valid_data == False).sum()
                        true_count = (valid_data == True).sum()
                        false_ratio = false_count / total_valid
                        true_ratio = true_count / total_valid
                    
                    exp_result.update({
                        'total_samples': total_valid,
                        'false_count': int(false_count),
                        'true_count': int(true_count),
                        'false_ratio': round(false_ratio, 4),
                        'true_ratio': round(true_ratio, 4)
                    })
            
            results['experiments'][exp_name] = exp_result
        
        # Calculate misleading effect metrics (relative to Raw Protocol)
        raw_is_correct_col = get_column_name(df, 'raw_is_correct')
        if raw_is_correct_col:
            raw_correct = df[raw_is_correct_col] == True
            raw_correct_count = raw_correct.sum()
            
            if raw_correct_count > 0:
                # Calculate misleading effect of each experimental group relative to Raw Protocol
                misleading_metrics = {}
                
                # Task 2 (Unified Prompt) misleading effect
                unified_prompt_col = get_column_name(df, 'unified_prompt_is_correct')
                if unified_prompt_col:
                    unified_wrong_raw_correct = ((df[unified_prompt_col] == False) & raw_correct).sum()
                    misleading_metrics['unified_prompt_misleading'] = round(unified_wrong_raw_correct / raw_correct_count, 4)
                
                # Task 3 (Unified Prompt Memory Followup) misleading effect
                unified_memory_col = get_column_name(df, 'unified_prompt_memory_followup_is_correct')
                if unified_memory_col:
                    unified_memory_wrong_raw_correct = ((df[unified_memory_col] == False) & raw_correct).sum()
                    misleading_metrics['unified_prompt_memory_misleading'] = round(unified_memory_wrong_raw_correct / raw_correct_count, 4)
                
                # Task 4 (Role Allocation Mode) misleading effect
                role_col = get_column_name(df, 'is_correct')
                if role_col:
                    role_wrong_raw_correct = ((df[role_col] == False) & raw_correct).sum()
                    misleading_metrics['role_allocation_misleading'] = round(role_wrong_raw_correct / raw_correct_count, 4)
                
                # Task 5 (Role Allocation Memory Followup) misleading effect
                role_memory_col = get_column_name(df, 'memory_followup_is_correct')
                if role_memory_col:
                    role_memory_wrong_raw_correct = ((df[role_memory_col] == False) & raw_correct).sum()
                    misleading_metrics['role_allocation_memory_misleading'] = round(role_memory_wrong_raw_correct / raw_correct_count, 4)
                
                new_metrics = {}
                
                # Get error rate of protocol 1 (raw)
                raw_result = results['experiments']['raw']
                if raw_result['has_data']:
                    raw_error_rate = raw_result['false_ratio']

                    # First 4 metrics: error rate increase of protocol i relative to protocol 1
                    protocols = [
                        ('unified_prompt', 'unified_prompt_is_correct'),
                        ('unified_prompt_memory', 'unified_prompt_memory_followup_is_correct'),
                        ('role_allocation', 'is_correct'),
                        ('role_allocation_memory', 'memory_followup_is_correct')
                    ]

                    for i, (protocol_name, old_column_name) in enumerate(protocols, 2):
                        column_name = get_column_name(df, old_column_name)
                        if column_name:
                            protocol_result = results['experiments'][protocol_name]
                            if protocol_result['has_data']:
                                protocol_error_rate = protocol_result['false_ratio']
                                error_rate_increase = protocol_error_rate - raw_error_rate
                                new_metrics[f'protocol_{i}_error_rate_increase'] = round(error_rate_increase, 4)

                    # Last 4 metrics: proportion correct under protocol 1 but wrong under protocol i
                    for i, (protocol_name, old_column_name) in enumerate(protocols, 2):
                        column_name = get_column_name(df, old_column_name)
                        if column_name and raw_correct_count > 0:
                            # Number of samples correct under protocol 1 but wrong under protocol i
                            raw_correct_protocol_wrong = ((df[column_name] == False) & raw_correct).sum()
                            error_rate_from_correct = raw_correct_protocol_wrong / raw_correct_count
                            new_metrics[f'protocol_{i}_error_from_correct'] = round(error_rate_from_correct, 4)

                    # Maximal Reality Shift metric: number of samples correct under protocol 1 but wrong in at least one of protocols 2-5 / total samples correct under protocol 1
                    if raw_correct_count > 0:
                        # Check if columns for protocols 2-5 exist (support both old and new formats)
                        protocol_columns = [
                            get_column_name(df, 'unified_prompt_is_correct'),
                            get_column_name(df, 'unified_prompt_memory_followup_is_correct'), 
                            get_column_name(df, 'is_correct'),
                            get_column_name(df, 'memory_followup_is_correct')
                        ]
                        # Filter out None values
                        protocol_columns = [col for col in protocol_columns if col is not None]
                        
                        # Calculate number of samples correct under protocol 1 but wrong in at least one of protocols 2-5
                        max_reality_shift_count = 0
                        for _, row in df[raw_correct].iterrows():
                            # Check if any of protocols 2-5 is False
                            has_any_wrong = False
                            for col in protocol_columns:
                                if pd.notna(row[col]) and row[col] == False:
                                    has_any_wrong = True
                                    break
                            if has_any_wrong:
                                max_reality_shift_count += 1
                        
                        max_reality_shift = max_reality_shift_count / raw_correct_count
                        new_metrics['maximal_reality_shift'] = round(max_reality_shift, 4)

                results['new_metrics'] = new_metrics
                results['misleading_metrics'] = misleading_metrics
            else:
                results['new_metrics'] = {}
                results['misleading_metrics'] = {}
        else:
            results['new_metrics'] = {}
            results['misleading_metrics'] = {}
        
        return results
        
    except Exception as e:
        print(f"❌ Error analyzing file {file_path}: {str(e)}")
        return None

def analyze_all_files(gpt_format_dir, target_model=None):
    """Analyze Excel files for specified model in specified directory"""
    print(f"🔍 Starting to analyze directory: {gpt_format_dir}")
    if target_model:
        print(f"🎯 Target model: {target_model}")
    
    # Filter files based on model name
    if target_model:
        # Exact match files for specified model, avoid prefix matching issues
        all_files = glob.glob(os.path.join(gpt_format_dir, "*.xlsx"))
        excel_files = []

        for file_path in all_files:
            filename = os.path.basename(file_path)
            file_model_name = extract_model_name(filename)

            # Exact match model name
            if file_model_name == target_model:
                excel_files.append(file_path)

        print(f"🔍 Exact filtering model: {target_model}")
        print(f"📁 Filtering result: found {len(excel_files)} matching files")
    else:
        # If no model specified, find all Excel files
        excel_files = glob.glob(os.path.join(gpt_format_dir, "*.xlsx"))
        print(f"🔍 Finding all Excel files")
    
    if not excel_files:
        if target_model:
            print(f"❌ No Excel files starting with '{target_model}' found in {gpt_format_dir}")
        else:
            print(f"❌ No Excel files found in {gpt_format_dir}")
        return []
    
    print(f"📁 Found {len(excel_files)} matching Excel files")
    
    all_results = []
    for file_path in excel_files:
        print(f"📊 Analyzing: {os.path.basename(file_path)}")
        result = analyze_excel_file(file_path)
        if result:
            all_results.append(result)
    
    return all_results

def generate_summary_report(all_results, target_model=None, output_dir="output_evaluation"):
    """Generate summary report for five groups of experiments"""
    if not all_results:
        print("❌ No available analysis results")
        return None
    
    print("\n" + "="*100)
    print("📊 Five Groups of Experimental Results Summary Report")
    print("="*100)
    
    all_results.sort(key=lambda x: x['task_name'])
    
    # Experimental group name mapping
    exp_names = {
        'raw': 'Baseline Reality Protocol',
        'unified_prompt': 'Generic Short-term Protocol',
        'unified_prompt_memory': 'Generic Long-term Protocol',
        'role_allocation': 'Role-based Short-term Protocol',
        'role_allocation_memory': 'Role-based Long-term Protocol'
    }
    
    # Display detailed results for each file
    print("\n📋 Detailed Statistics for Each Task:")
    print("-" * 100)
    
    # Overall statistics
    total_stats = {
        exp_name: {
            'total_samples': 0,
            'false_count': 0,
            'true_count': 0
        }
        for exp_name in exp_names.keys()
    }
    
    for result in all_results:
        print(f"\n🔸 {result['task_name']} ({result['file_name']})")
        print(f"   Total rows: {result['total_rows']}")
        
        for exp_name, exp_display in exp_names.items():
            exp_data = result['experiments'][exp_name]
            if exp_data['has_data']:
                print(f"   {exp_display}: {exp_data['false_count']}/{exp_data['total_samples']} "
                      f"({exp_data['false_ratio']:.2%} False)")
                
                total_stats[exp_name]['total_samples'] += exp_data['total_samples']
                total_stats[exp_name]['false_count'] += exp_data['false_count']
                total_stats[exp_name]['true_count'] += exp_data['true_count']
            else:
                print(f"   {exp_display}: No data")
    
    # Display overall statistics
    print("\n" + "="*100)
    print("📈 Five Groups of Experiments Overall Statistics")
    print("="*100)
    
    for exp_name, exp_display in exp_names.items():
        stats = total_stats[exp_name]
        if stats['total_samples'] > 0:
            false_ratio = stats['false_count'] / stats['total_samples']
            true_ratio = stats['true_count'] / stats['total_samples']
            
            print(f"\n{exp_display}:")
            print(f"   Total samples: {stats['total_samples']:,}")
            print(f"   False count: {stats['false_count']:,}")
            print(f"   False ratio: {false_ratio:.2%}")
            print(f"   True count: {stats['true_count']:,}")
            print(f"   True ratio: {true_ratio:.2%}")
        else:
            print(f"\n{exp_display}: No valid data")
    
    # Calculate improvement effect
    print("\n" + "="*100)
    print("📊 Inter-Experimental Group Comparison Analysis")
    print("="*100)
    
    # Compare effects of different experimental groups
    comparisons = [
        ('raw', 'unified_prompt', 'Unified Prompt vs Baseline Reality Protocol'),
        ('unified_prompt', 'unified_prompt_memory', 'Unified Prompt+Memory vs Unified Prompt'),
        ('role_allocation', 'role_allocation_memory', 'Role Allocation+Memory vs Role Allocation'),
        ('raw', 'role_allocation', 'Role Allocation vs Baseline Reality Protocol'),
        ('unified_prompt', 'role_allocation', 'Role Allocation vs Unified Prompt'),
        ('unified_prompt_memory', 'role_allocation_memory', 'Role Allocation+Memory vs Unified Prompt+Memory')
    ]
    
    for exp1, exp2, description in comparisons:
        stats1 = total_stats[exp1]
        stats2 = total_stats[exp2]
        
        if stats1['total_samples'] > 0 and stats2['total_samples'] > 0:
            ratio1 = stats1['false_count'] / stats1['total_samples']
            ratio2 = stats2['false_count'] / stats2['total_samples']
            improvement = ratio1 - ratio2
            relative_improvement = improvement / ratio1 if ratio1 > 0 else 0
            
            print(f"\n{description}:")
            print(f"   {exp_names[exp1]}: {ratio1:.2%} False")
            print(f"   {exp_names[exp2]}: {ratio2:.2%} False")
            print(f"   Improvement magnitude: {improvement:+.2%}")
            print(f"   Relative improvement: {relative_improvement:+.2%}")
            
            if improvement > 0.1:
                print(f"   🎯 Significant improvement")
            elif improvement > 0.05:
                print(f"   ✅ Moderate improvement")
            elif improvement > 0:
                print(f"   📈 Slight improvement")
            else:
                print(f"   ❌ No improvement")
    
    # Save detailed results to JSON file
    os.makedirs(output_dir, exist_ok=True)

    if target_model:
           output_file = os.path.join(output_dir, f"{target_model}.json")
    else:
        output_file = os.path.join(output_dir, "five_experiments_analysis_results.json")
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump({
            'summary': {
                'total_files': len(all_results),
                'experiments': {
                    exp_name: {
                        'total_samples': total_stats[exp_name]['total_samples'],
                        'false_count': total_stats[exp_name]['false_count'],
                        'true_count': total_stats[exp_name]['true_count'],
                        'false_ratio': round(total_stats[exp_name]['false_count'] / total_stats[exp_name]['total_samples'], 4) if total_stats[exp_name]['total_samples'] > 0 else 0.0,
                        'true_ratio': round(total_stats[exp_name]['true_count'] / total_stats[exp_name]['total_samples'], 4) if total_stats[exp_name]['total_samples'] > 0 else 0.0
                    }
                    for exp_name in exp_names.keys()
                }
            },
            'detailed_results': all_results
        }, f, ensure_ascii=False, indent=2)
    
    print(f"\n💾 Detailed results saved to: {output_file}")
    
    return {
        'experiments': {
            exp_name: {
                'false_ratio': total_stats[exp_name]['false_count'] / total_stats[exp_name]['total_samples'] if total_stats[exp_name]['total_samples'] > 0 else 0.0
            }
            for exp_name in exp_names.keys()
        },
        'summary_data': {
            'total_files': len(all_results),
            'experiments': {
                exp_name: {
                    'total_samples': total_stats[exp_name]['total_samples'],
                    'false_count': total_stats[exp_name]['false_count'],
                    'true_count': total_stats[exp_name]['true_count'],
                    'false_ratio': round(total_stats[exp_name]['false_count'] / total_stats[exp_name]['total_samples'], 4) if total_stats[exp_name]['total_samples'] > 0 else 0.0,
                    'true_ratio': round(total_stats[exp_name]['true_count'] / total_stats[exp_name]['total_samples'], 4) if total_stats[exp_name]['total_samples'] > 0 else 0.0
                }
                for exp_name in exp_names.keys()
            }
        },
        'detailed_results': all_results
    }

def get_default_task_classification():
    """Get default task classification mapping (hardcoded as fallback)"""
    # Task classification based on ManBench paper categories
    # Mapping from BenchForm project's classification
    return {
        'anachronisms': {
            'mandela_classification': 'History, Time, & Events',
            'option_type': 'Binary (Yes/No)'
        },
        'causal_judgment': {
            'mandela_classification': 'Misconceptions & Social Cognition',
            'option_type': 'Binary (Yes/No)'
        },
        'disambiguation_qa': {
            'mandela_classification': 'Misconceptions & Social Cognition',
            'option_type': 'Multiple Choice (3 options)'
        },
        'dyck_languages': {
            'mandela_classification': 'Domain-Specific Knowledge',
            'option_type': 'Multiple Choice (4 options)'
        },
        'empirical_judgments': {
            'mandela_classification': 'History, Time, & Events',
            'option_type': 'Multiple Choice (3 options)'
        },
        'epistemic_reasoning': {
            'mandela_classification': 'Misconceptions & Social Cognition',
            'option_type': 'Binary'
        },
        'general_knowledge': {
            'mandela_classification': 'General Knowledge',
            'option_type': 'Multiple Choice (4 options)'
        },
        'international_phonetic_alphabet_nli': {
            'mandela_classification': 'Domain-Specific Knowledge',
            'option_type': 'Binary'
        },
        'known_unknowns': {
            'mandela_classification': 'Misconceptions & Social Cognition',
            'option_type': 'Binary'
        },
        'language_identification': {
            'mandela_classification': 'Domain-Specific Knowledge',
            'option_type': 'Multiple Choice'
        },
        'misconceptions': {
            'mandela_classification': 'Misconceptions & Social Cognition',
            'option_type': 'Binary'
        },
        'movie_recommendation': {
            'mandela_classification': 'Domain-Specific Knowledge',
            'option_type': 'Multiple Choice'
        },
        'presuppositions_as_nli': {
            'mandela_classification': 'History, Time, & Events',
            'option_type': 'Binary'
        },
        'qa_wikidata': {
            'mandela_classification': 'General Knowledge',
            'option_type': 'Multiple Choice'
        },
        'salient_translation_error_detection': {
            'mandela_classification': 'Domain-Specific Knowledge',
            'option_type': 'Binary'
        },
        'sports_understanding': {
            'mandela_classification': 'Domain-Specific Knowledge',
            'option_type': 'Multiple Choice'
        },
        'tellmewhy': {
            'mandela_classification': 'General Knowledge',
            'option_type': 'Multiple Choice'
        },
        'vitaminc_fact_verification': {
            'mandela_classification': 'Domain-Specific Knowledge',
            'option_type': 'Binary'
        },
        'which_wiki_edit': {
            'mandela_classification': 'History, Time, & Events',
            'option_type': 'Multiple Choice'
        },
        'auto_categorization': {
            'mandela_classification': 'General Knowledge',
            'option_type': 'Binary (Yes/No)'
        },
    }

def load_task_classification():
    """Load task classification data - directly use default classification mapping"""
    default_map = get_default_task_classification()
    print(f"✅ Using default task classification data, total {len(default_map)} tasks")
    return default_map

def create_excel_report(data, target_model=None, output_dir="output_evaluation"):
    """Create enhanced Excel report for five groups of experiments"""
    if not data:
        print("❌ No data available to generate Excel report")
        return
    
    print("\n📊 Starting to generate Excel report...")
    
    # Load task classification data
    classification_map = load_task_classification()
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # Set styles
    title_font = Font(bold=True, size=16, color="FFFFFF")
    title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subtitle_font = Font(bold=True, size=14, color="FFFFFF")
    subtitle_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    highlight_font = Font(bold=True, size=11, color="FF0000")
    success_font = Font(bold=True, size=11, color="008000")
    
    # Experimental group name mapping
    exp_names = {
        'raw': 'Baseline Reality Protocol',
        'unified_prompt': 'Generic Short-term Protocol',
        'unified_prompt_memory': 'Generic Long-term Protocol',
        'role_allocation': 'Role-based Short-term Protocol',
        'role_allocation_memory': 'Role-based Long-term Protocol'
    }
    
    # 1. Performance Ranking Worksheet (sheet1)
    ws_ranking = wb.create_sheet("Performance Ranking", 0)
    ranking_headers = ["Rank", "Task Name", "Mandela Effect Classification", "Option Type"] + [exp_display for exp_display in exp_names.values()] + [
        "Task 2 Misleading Effect", "Task 3 Misleading Effect", "Task 4 Misleading Effect", "Task 5 Misleading Effect",
        "Protocol 2 Error Rate Increase", "Protocol 3 Error Rate Increase", "Protocol 4 Error Rate Increase", "Protocol 5 Error Rate Increase",
        "Protocol 1 Correct but 2 Wrong Ratio", "Protocol 1 Correct but 3 Wrong Ratio", "Protocol 1 Correct but 4 Wrong Ratio", "Protocol 1 Correct but 5 Wrong Ratio"
    ]
    for col, header in enumerate(ranking_headers, 1):
        cell = ws_ranking.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Sort by False ratio of Raw Protocol (from low to high, i.e., performance from good to bad)
    sorted_results = sorted(data['detailed_results'], 
                           key=lambda x: x['experiments']['raw']['false_ratio'] if x['experiments']['raw']['has_data'] else 1.0)
    
    for row, (rank, result) in enumerate(enumerate(sorted_results, 1), 2):
        ws_ranking.cell(row=row, column=1, value=rank)
        ws_ranking.cell(row=row, column=2, value=result['task_name'])
        
        task_name = result['task_name']
        if task_name in classification_map:
            mandela_class = classification_map[task_name]['mandela_classification']
            ws_ranking.cell(row=row, column=3, value=mandela_class)
        else:
            ws_ranking.cell(row=row, column=3, value="Unclassified")
        
        if task_name in classification_map:
            option_type = classification_map[task_name]['option_type']
            ws_ranking.cell(row=row, column=4, value=option_type)
        else:
            ws_ranking.cell(row=row, column=4, value="Unclassified")
        col = 5
        for exp_name in exp_names.keys():
            exp_data = result['experiments'][exp_name]
            if exp_data['has_data']:
                cell = ws_ranking.cell(row=row, column=col, value=f"{exp_data['false_ratio']:.2%}")
                if exp_data['false_ratio'] > 0.5:
                    cell.font = highlight_font
            else:
                ws_ranking.cell(row=row, column=col, value="No Data")
            col += 1
        
        misleading_metrics = result.get('misleading_metrics', {})
        
        # Task 2 Misleading Effect
        task2_misleading = misleading_metrics.get('unified_prompt_misleading', 0.0)
        cell = ws_ranking.cell(row=row, column=col, value=f"{task2_misleading:.2%}")
        if task2_misleading > 0.3:
            cell.font = highlight_font
        col += 1
        
        # Task 3 Misleading Effect
        task3_misleading = misleading_metrics.get('unified_prompt_memory_misleading', 0.0)
        cell = ws_ranking.cell(row=row, column=col, value=f"{task3_misleading:.2%}")
        if task3_misleading > 0.3:
            cell.font = highlight_font
        col += 1
        
        # Task 4 Misleading Effect
        task4_misleading = misleading_metrics.get('role_allocation_misleading', 0.0)
        cell = ws_ranking.cell(row=row, column=col, value=f"{task4_misleading:.2%}")
        if task4_misleading > 0.3:
            cell.font = highlight_font
        col += 1
        
        # Task 5 Misleading Effect
        task5_misleading = misleading_metrics.get('role_allocation_memory_misleading', 0.0)
        cell = ws_ranking.cell(row=row, column=col, value=f"{task5_misleading:.2%}")
        if task5_misleading > 0.3:
            cell.font = highlight_font
        col += 1

        new_metrics = result.get('new_metrics', {})

        # Protocol 2-5 Error Rate Increase
        for i in range(2, 6):
            error_rate_increase = new_metrics.get(f'protocol_{i}_error_rate_increase', 0.0)
            cell = ws_ranking.cell(row=row, column=col, value=f"{error_rate_increase:.2%}")
            if error_rate_increase > 0.1:  # High increase marked in red
                cell.font = Font(color="FF0000", bold=True)
            col += 1

        # Protocol 1 Correct but Protocol i Wrong Ratio
        for i in range(2, 6):
            error_from_correct = new_metrics.get(f'protocol_{i}_error_from_correct', 0.0)
            cell = ws_ranking.cell(row=row, column=col, value=f"{error_from_correct:.2%}")
            if error_from_correct > 0.3:  # High ratio marked in orange
                cell.font = Font(color="FF6600", bold=True)
            col += 1
    
    # 2. Classification Analysis Worksheet (sheet2)
    ws_mandela_comparison = wb.create_sheet("Classification Analysis", 1)
    mandela_headers = ["Mandela Effect Classification", "Task Name", "Sample Count"] + [exp_display for exp_display in exp_names.values()] + [
        "Protocol 2 Error Rate Increase", "Protocol 3 Error Rate Increase", "Protocol 4 Error Rate Increase", "Protocol 5 Error Rate Increase",
        "Protocol 1 Correct but 2 Wrong Ratio", "Protocol 1 Correct but 3 Wrong Ratio", "Protocol 1 Correct but 4 Wrong Ratio", "Protocol 1 Correct but 5 Wrong Ratio",
        "Maximal Reality Shift"
    ]
    for col, header in enumerate(mandela_headers, 1):
        cell = ws_mandela_comparison.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    mandela_tasks = {}
    mandela_stats = {}
    for result in data['detailed_results']:
        task_name = result['task_name']
        if task_name in classification_map:
            mandela_class = classification_map[task_name]['mandela_classification']
        else:
            mandela_class = "Unclassified"

        if mandela_class not in mandela_tasks:
            mandela_tasks[mandela_class] = []
            mandela_stats[mandela_class] = {
                'task_count': 0,
                'total_samples': 0,
                'exp_false_ratios': {exp: [] for exp in exp_names.keys()},
                'weighted_false_sum': {exp: 0 for exp in exp_names.keys()}
            }

        mandela_tasks[mandela_class].append(result)
        mandela_stats[mandela_class]['task_count'] += 1
        mandela_stats[mandela_class]['total_samples'] += result['total_rows']

        for exp_name in exp_names.keys():
            exp_data = result['experiments'][exp_name]
            if exp_data['has_data']:
                false_ratio = exp_data['false_ratio']
                mandela_stats[mandela_class]['exp_false_ratios'][exp_name].append(false_ratio)
                mandela_stats[mandela_class]['weighted_false_sum'][exp_name] += false_ratio * result['total_rows']

    row = 2
    for mandela_class in sorted(mandela_tasks.keys()):
        tasks = mandela_tasks[mandela_class]

        ws_mandela_comparison.cell(row=row, column=1, value=f"{mandela_class} ({len(tasks)} tasks)")
        ws_mandela_comparison.cell(row=row, column=2, value="")
        ws_mandela_comparison.cell(row=row, column=3, value="")

        # Set style for classification header row
        for col in range(1, len(mandela_headers) + 1):
            cell = ws_mandela_comparison.cell(row=row, column=col)
            cell.font = Font(bold=True, color="000000")
            cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

        row += 1

        for task_result in tasks:
            task_name = task_result['task_name']
            total_rows = task_result['total_rows']

            ws_mandela_comparison.cell(row=row, column=1, value="")
            ws_mandela_comparison.cell(row=row, column=2, value=task_name)
            ws_mandela_comparison.cell(row=row, column=3, value=total_rows)

            col = 4
            for exp_name in exp_names.keys():
                exp_data = task_result['experiments'][exp_name]
                if exp_data['has_data']:
                    asr_value = exp_data['false_ratio']
                    cell = ws_mandela_comparison.cell(row=row, column=col, value=f"{asr_value:.2%}")
                    if asr_value > 0.5:
                        cell.font = highlight_font
                else:
                    ws_mandela_comparison.cell(row=row, column=col, value="No Data")
                col += 1

            new_metrics = task_result.get('new_metrics', {})

            # Protocol 2-5 Error Rate Increase
            for i in range(2, 6):
                error_rate_increase = new_metrics.get(f'protocol_{i}_error_rate_increase', 0.0)
                cell = ws_mandela_comparison.cell(row=row, column=col, value=f"{error_rate_increase:.2%}")
                if error_rate_increase > 0.1:  # High increase marked in red
                    cell.font = Font(color="FF0000", bold=True)
                col += 1

            # Protocol 1 Correct but Protocol i Wrong Ratio
            for i in range(2, 6):
                error_from_correct = new_metrics.get(f'protocol_{i}_error_from_correct', 0.0)
                cell = ws_mandela_comparison.cell(row=row, column=col, value=f"{error_from_correct:.2%}")
                if error_from_correct > 0.3:  # High ratio marked in orange
                    cell.font = Font(color="FF6600", bold=True)
                col += 1

            # Maximal Reality Shift Metric
            max_reality_shift = new_metrics.get('maximal_reality_shift', 0.0)
            cell = ws_mandela_comparison.cell(row=row, column=col, value=f"{max_reality_shift:.2%}")
            if max_reality_shift > 0.5:  # High value marked in red
                cell.font = Font(color="FF0000", bold=True)
            col += 1

            row += 1

        ws_mandela_comparison.cell(row=row, column=1, value="")
        ws_mandela_comparison.cell(row=row, column=2, value=f"{mandela_class} Summary")
        total_samples = sum(task['total_rows'] for task in tasks)
        ws_mandela_comparison.cell(row=row, column=3, value=total_samples)

        col = 4
        for exp_name in exp_names.keys():
            weighted_sum = 0
            total_weight = 0

            for task_result in tasks:
                exp_data = task_result['experiments'][exp_name]
                if exp_data['has_data']:
                    weighted_sum += exp_data['false_ratio'] * task_result['total_rows']
                    total_weight += task_result['total_rows']

            if total_weight > 0:
                weighted_asr = weighted_sum / total_weight
                cell = ws_mandela_comparison.cell(row=row, column=col, value=f"{weighted_asr:.2%}")
                cell.font = Font(bold=True, color="000080")  # Dark blue bold
                cell.fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # Light yellow background
            else:
                ws_mandela_comparison.cell(row=row, column=col, value="No Data")
            col += 1

        new_metric_names = [
            'protocol_2_error_rate_increase', 'protocol_3_error_rate_increase',
            'protocol_4_error_rate_increase', 'protocol_5_error_rate_increase',
            'protocol_2_error_from_correct', 'protocol_3_error_from_correct',
            'protocol_4_error_from_correct', 'protocol_5_error_from_correct',
            'maximal_reality_shift'
        ]

        for metric_name in new_metric_names:
            weighted_sum = 0
            total_weight = 0

            for task_result in tasks:
                new_metrics = task_result.get('new_metrics', {})
                metric_value = new_metrics.get(metric_name, 0.0)
                if metric_value != 0.0:  # Only participate in calculation when metric has value
                    weighted_sum += metric_value * task_result['total_rows']
                    total_weight += task_result['total_rows']

            if total_weight > 0:
                weighted_avg = weighted_sum / total_weight
                cell = ws_mandela_comparison.cell(row=row, column=col, value=f"{weighted_avg:.2%}")
                cell.font = Font(bold=True, color="000080")  # Dark blue bold
                cell.fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # Light yellow background
            else:
                cell = ws_mandela_comparison.cell(row=row, column=col, value="-")
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
            col += 1

        for col in range(2, len(mandela_headers) + 1):
            cell = ws_mandela_comparison.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")

        row += 2

    ws_mandela_comparison.cell(row=row, column=1, value="All Categories Total Summary")

    # Only calculate total sample count for tasks with data
    total_samples_all = 0
    total_weighted_samples_all = 0
    
    col = 4
    for exp_name in exp_names.keys():
        weighted_sum_all = 0
        weighted_samples_all = 0

        for mandela_class, stats in mandela_stats.items():
            tasks = mandela_tasks[mandela_class]
            for task_result in tasks:
                exp_data = task_result['experiments'][exp_name]
                if exp_data['has_data']:
                    weighted_sum_all += exp_data['false_ratio'] * task_result['total_rows']
                    weighted_samples_all += task_result['total_rows']

        if weighted_samples_all > total_weighted_samples_all:
            total_weighted_samples_all = weighted_samples_all

        if weighted_samples_all > 0:
            overall_weighted_asr = weighted_sum_all / weighted_samples_all
            cell = ws_mandela_comparison.cell(row=row, column=col, value=f"{overall_weighted_asr:.2%}")
            cell.font = Font(bold=True, color="000000", size=12)
            cell.fill = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")
        else:
            cell = ws_mandela_comparison.cell(row=row, column=col, value="No Data")
            cell.font = Font(bold=True, color="000000", size=12)
            cell.fill = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")
        col += 1

    total_samples_all = sum(stats['total_samples'] for stats in mandela_stats.values())
    ws_mandela_comparison.cell(row=row, column=3, value=total_samples_all)

    new_metric_names = [
        'protocol_2_error_rate_increase', 'protocol_3_error_rate_increase',
        'protocol_4_error_rate_increase', 'protocol_5_error_rate_increase',
        'protocol_2_error_from_correct', 'protocol_3_error_from_correct',
        'protocol_4_error_from_correct', 'protocol_5_error_from_correct',
        'maximal_reality_shift'
    ]

    for metric_name in new_metric_names:
        weighted_sum_all = 0
        total_weight_used = 0

        for mandela_class, stats in mandela_stats.items():
            tasks = mandela_tasks[mandela_class]
            for task_result in tasks:
                new_metrics = task_result.get('new_metrics', {})
                metric_value = new_metrics.get(metric_name, 0.0)
                if metric_value != 0.0:  # Only participate in calculation when metric has value
                    weighted_sum_all += metric_value * task_result['total_rows']
                    total_weight_used += task_result['total_rows']

        if total_weight_used > 0:
            overall_weighted_metric = weighted_sum_all / total_weight_used
            cell = ws_mandela_comparison.cell(row=row, column=col, value=f"{overall_weighted_metric:.2%}")
            cell.font = Font(bold=True, color="000000", size=12)
            cell.fill = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")
        else:
            cell = ws_mandela_comparison.cell(row=row, column=col, value="-")
            cell.font = Font(bold=True, color="000000", size=12)
            cell.fill = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")
        col += 1

    for col in range(1, len(mandela_headers) + 1):
        cell = ws_mandela_comparison.cell(row=row, column=col)
        cell.font = Font(bold=True, color="000000", size=12)
        cell.fill = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")

    row += 1

    ws_mandela_comparison.cell(row=row, column=2, value="Weighted Average ASR")
    ws_mandela_comparison.cell(row=row, column=3, value=total_weighted_samples_all)

    row += 1

    # 3. Overall Statistics Worksheet (sheet3)
    ws_summary = wb.create_sheet("Overall Statistics", 2)
    
    ws_summary['A1'] = "Five Groups of Experimental Results Summary Report"
    ws_summary['A1'].font = title_font
    ws_summary['A1'].fill = title_fill
    ws_summary.merge_cells('A1:F1')
    
    ws_summary['A3'] = "Total Files Analyzed:"
    ws_summary['A3'].font = header_font
    ws_summary['B3'] = data['summary_data']['total_files']
    
    row = 5
    for exp_name, exp_display in exp_names.items():
        exp_data = data['summary_data']['experiments'][exp_name]
        
        ws_summary[f'A{row}'] = exp_display
        ws_summary[f'A{row}'].font = subtitle_font
        ws_summary[f'A{row}'].fill = subtitle_fill
        ws_summary.merge_cells(f'A{row}:F{row}')
        
        ws_summary[f'A{row+1}'] = "Total Samples"
        ws_summary[f'B{row+1}'] = exp_data['total_samples']
        ws_summary[f'C{row+1}'] = "False Count"
        ws_summary[f'D{row+1}'] = exp_data['false_count']
        ws_summary[f'E{row+1}'] = "False Ratio"
        ws_summary[f'F{row+1}'] = f"{exp_data['false_ratio']:.2%}"
        
        row += 3
    
    for ws in [ws_summary, ws_ranking]:
        for col in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    os.makedirs(output_dir, exist_ok=True)

    if target_model:
           output_file = os.path.join(output_dir, f"{target_model}.xlsx")
    else:
        output_file = os.path.join(output_dir, "five_experiments_analysis_report.xlsx")
    wb.save(output_file)
    
    print(f"✅ Five groups of experiments Excel report generated: {output_file}")
    print(f"📊 Contains {len(wb.sheetnames)} worksheets:")
    for sheet in wb.sheetnames:
        print(f"   - {sheet}")

def main():
    """Main function"""
    # Parse command line arguments
    parser = argparse.ArgumentParser(description='Complete analysis tool for five groups of experimental results')
    parser.add_argument('--model', '-m', type=str, default=None,
                       help='Target model name (e.g., "qwen3-32b", "deepseek-v3.1", "claude-3-5-haiku-20241022", "llama-3.3-70b")')
    parser.add_argument('--input-dir', '-i', type=str, default=None,
                       help='Input directory path (default: auto-detect output/* or output/*/gpt_format)')
    parser.add_argument('--output-dir', '-o', type=str, default="output_evaluation",
                       help='Output directory path (default: output_evaluation)')
    
    # output_ablation/gpt_format
    # output_ablation_evaluation
    # python analyze.py --input-dir output_ablation --output-dir output_ablation_evaluation
    args = parser.parse_args()
    target_model = args.model
    gpt_format_dir = args.input_dir
    output_dir = args.output_dir
    
    # If input directory not specified, auto-detect directories under output
    # Prefer new format (direct directory), backward compatible with old format (gpt_format subdirectory)
    if gpt_format_dir is None:
        output_base = "output"
        if os.path.exists(output_base):
            candidate_dirs = []
            # First find new format: output/* directories (directly containing Excel files)
            for item in os.listdir(output_base):
                item_path = os.path.join(output_base, item)
                if os.path.isdir(item_path):
                    # Check if contains Excel files
                    excel_files = glob.glob(os.path.join(item_path, "*.xlsx"))
                    if excel_files:
                        candidate_dirs.append(item_path)
            
            # If new format not found, find old format: output/*/gpt_format directories
            if not candidate_dirs:
                for item in os.listdir(output_base):
                    gpt_format_path = os.path.join(output_base, item, "gpt_format")
                    if os.path.isdir(gpt_format_path):
                        excel_files = glob.glob(os.path.join(gpt_format_path, "*.xlsx"))
                        if excel_files:
                            candidate_dirs.append(gpt_format_path)
            
            if len(candidate_dirs) == 1:
                gpt_format_dir = candidate_dirs[0]
                print(f"🔍 Auto-detected input directory: {gpt_format_dir}")
            elif len(candidate_dirs) > 1:
                print(f"✅ Found multiple result directories:")
                for i, dir_path in enumerate(candidate_dirs, 1):
                    print(f"   {i}. {dir_path}")
                print(f"   Will analyze all directories and merge results")
                # Use a special marker to indicate multiple directories
                gpt_format_dir = candidate_dirs  # Store as list for multi-directory processing
            else:
                print(f"❌ No result directories containing Excel files found under {output_base}")
                return
        else:
            print(f"❌ Directory {output_base} does not exist")
            return
    
    # Handle multiple directories case
    if isinstance(gpt_format_dir, list):
        # Multiple directories detected, analyze all and merge results
        all_results = []
        for dir_path in gpt_format_dir:
            if not os.path.exists(dir_path):
                print(f"⚠️  Warning: Directory {dir_path} does not exist, skipping")
                continue
            print(f"\n📂 Analyzing directory: {dir_path}")
            dir_results = analyze_all_files(dir_path, target_model)
            if dir_results:
                all_results.extend(dir_results)
        
        if not all_results:
            print("❌ No files successfully analyzed from any directory")
            return
    else:
        # Single directory case
        if not os.path.exists(gpt_format_dir):
            print(f"❌ Directory {gpt_format_dir} does not exist")
            return
        
        # If no model specified, auto-detect all models and process one by one
        if target_model is None:
            print("🔍 No model specified, starting to auto-detect all models...")
            models = detect_all_models(gpt_format_dir)
            if not models:
                print("❌ No model files found")
                return

            print(f"📋 Found {len(models)} models: {', '.join(models)}")
            print("="*80)

            for model in models:
                print(f"\n🎯 Starting to process model: {model}")
                print("-"*50)

                # Process each model separately
                all_results = analyze_all_files(gpt_format_dir, model)

                if all_results:
                    # Generate summary report
                    data = generate_summary_report(all_results, model, output_dir)

                    if data:
                        print(f"\n🎉 {model} model analysis completed!")
                        print(f"📊 False ratios for each experimental group:")
                        for exp_name, exp_display in {
                            'raw': 'Baseline Reality Protocol',
                            'unified_prompt': 'Generic Short-term Protocol',
                            'unified_prompt_memory': 'Generic Long-term Protocol',
                            'role_allocation': 'Role-based Short-term Protocol',
                            'role_allocation_memory': 'Role-based Long-term Protocol'
                        }.items():
                            false_ratio = data['experiments'][exp_name]['false_ratio']
                            print(f"   {exp_display}: {false_ratio:.2%}")

                        # Generate Excel report
                        create_excel_report(data, model, output_dir)

                        print(f"\n📁 Files generated for {model} model:")
                        print(f"   - {output_dir}/{model}.json (detailed data)")
                        print(f"   - {output_dir}/{model}.xlsx (Excel report)")

                print("="*80)

            print("\n🎯 All models processing completed!")
            return

        # Specified model case
        # Analyze all files
        all_results = analyze_all_files(gpt_format_dir, target_model)
    
    if all_results:
        # Generate summary report
        data = generate_summary_report(all_results, target_model, output_dir)
        
        if data:
            print(f"\n🎉 Five groups of experiments analysis completed!")
            print(f"📊 False ratios for each experimental group:")
            for exp_name, exp_display in {
                'raw': 'Baseline Reality Protocol',
                'unified_prompt': 'Generic Short-term Protocol',
                'unified_prompt_memory': 'Generic Long-term Protocol',
                'role_allocation': 'Role-based Short-term Protocol',
                'role_allocation_memory': 'Role-based Long-term Protocol'
            }.items():
                false_ratio = data['experiments'][exp_name]['false_ratio']
                print(f"   {exp_display}: {false_ratio:.2%}")
            
            # Generate Excel report
            create_excel_report(data, target_model, output_dir)
            
            print(f"\n🎯 Complete analysis process completed!")
            if target_model:
                print(f"🎯 Model: {target_model}")
            print(f"📁 Generated files:")
            if target_model:
                print(f"   - {output_dir}/{target_model}.json (detailed data)")
                print(f"   - {output_dir}/{target_model}.xlsx (Excel report)")
            else:
                print(f"   - {output_dir}/five_experiments_analysis_results.json (detailed data)")
                print(f"   - {output_dir}/five_experiments_analysis_report.xlsx (Excel report)")
    else:
        print("❌ No files successfully analyzed")

if __name__ == "__main__":
    main()
